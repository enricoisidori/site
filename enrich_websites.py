#!/usr/bin/env python3
import argparse
import csv
import random
import re
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
import requests
from bs4 import BeautifulSoup, Comment
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/119.0 Safari/537.36"
)

REQUEST_TIMEOUT = 20
MAX_EXCERPT_CHARS = 1500
MAX_ABOUT_PAGES = 3


STOPWORDS = set(
    [
        "the",
        "and",
        "for",
        "with",
        "this",
        "that",
        "from",
        "have",
        "has",
        "are",
        "was",
        "were",
        "their",
        "they",
        "them",
        "our",
        "your",
        "you",
        "his",
        "her",
        "its",
        "about",
        "into",
        "over",
        "under",
        "within",
        "between",
        "across",
        "through",
        "studio",
        "works",
        "work",
        "based",
        "new",
        "york",
        "london",
        "paris",
        "berlin",
        "milan",
        "rome",
        "tokyo",
        "seoul",
        "china",
        "italy",
        "france",
        "usa",
        "united",
        "states",
        "project",
        "projects",
        "practice",
        "practices",
        "international",
        "based",
        "using",
        "use",
        "explores",
        "explore",
        "exploration",
        "research",
        "focus",
        "focusing",
        "focuses",
        "artist",
        "artists",
        "designer",
        "designers",
        "design",
        "art",
        "based",
        "studio",
        "company",
        "group",
        "collective",
        "team",
        "workshop",
        "about",
        "info",
        "bio",
        "statement",
        "cv",
        "email",
        "contact",
        "phone",
        "copyright",
        "all",
        "rights",
        "reserved",
    ]
)

WHITELIST_KEYWORDS = [
    # art / interdisciplinary leaning
    "installation",
    "performance",
    "research",
    "sound",
    "video",
    "sculpture",
    "painting",
    "media",
    "interactive",
    "interactivity",
    "generative",
    "spatial",
    "exhibition",
    "archive",
    "mapping",
    "data",
    "algorithm",
    "algorithmic",
    "sonic",
    "photography",
    "film",
    "cinema",
    "motion",
    "performance",
    "theatre",
    "theater",
    "projection",
    "virtual",
    "xr",
    "ar",
    "vr",
    # graphic leaning
    "typography",
    "type",
    "lettering",
    "editorial",
    "publication",
    "print",
    "poster",
    "branding",
    "identity",
    "logo",
    "packaging",
    "web",
    "website",
    "ui",
    "ux",
    "interface",
    "illustration",
    "infographics",
]

GRAPHIC_KEYWORDS = set(
    [
        "typography",
        "type",
        "lettering",
        "editorial",
        "publication",
        "print",
        "poster",
        "branding",
        "identity",
        "logo",
        "packaging",
        "web",
        "website",
        "ui",
        "ux",
        "interface",
        "illustration",
        "infographic",
        "infographics",
        "layout",
        "visual",
        "communication",
    ]
)

INTERDISCIPLINARY_HINTS = set(
    [
        "installation",
        "performance",
        "media",
        "research",
        "sound",
        "video",
        "sculpture",
        "painting",
        "interactive",
        "generative",
        "spatial",
        "exhibition",
        "mapping",
        "data",
        "photography",
        "motion",
        "projection",
    ]
)

EDU_TERMS = [
    "studied",
    "graduated",
    "ba",
    "ma",
    "mfa",
    "phd",
    "academy",
    "university",
]


def now_iso() -> str:
    return datetime.utcnow().isoformat(timespec="seconds") + "Z"


def normalize_url(u: Optional[str]) -> Optional[str]:
    if not u:
        return None
    u = u.strip()
    if not u:
        return None
    if not re.match(r"^https?://", u, re.I):
        u = "http://" + u
    return u


class RunLogger:
    def __init__(self, path: str):
        self.path = path
        self.fieldnames = [
            "timestamp",
            "name",
            "input_url",
            "action",
            "status",
            "page_url",
            "http_status",
            "bytes",
            "detail",
        ]
        self._init_file()

    def _init_file(self):
        try:
            with open(self.path, "x", newline="") as f:
                w = csv.DictWriter(f, fieldnames=self.fieldnames)
                w.writeheader()
        except FileExistsError:
            pass

    def log(
        self,
        name: str,
        input_url: Optional[str],
        action: str,
        status: str,
        page_url: Optional[str] = None,
        http_status: Optional[int] = None,
        size_bytes: Optional[int] = None,
        detail: Optional[str] = None,
    ):
        row = {
            "timestamp": now_iso(),
            "name": name or "",
            "input_url": input_url or "",
            "action": action,
            "status": status,
            "page_url": page_url or "",
            "http_status": http_status or "",
            "bytes": size_bytes or "",
            "detail": (detail or "")[:1000],
        }
        with open(self.path, "a", newline="") as f:
            w = csv.DictWriter(f, fieldnames=self.fieldnames)
            w.writerow(row)


class RateLimiter:
    def __init__(self, min_delay: float = 1.0, max_delay: float = 2.0):
        self.last_hit: Dict[str, float] = {}
        self.min_delay = min_delay
        self.max_delay = max_delay

    def wait(self, url: str):
        from urllib.parse import urlparse

        netloc = urlparse(url).netloc
        if not netloc:
            return
        now = time.time()
        last = self.last_hit.get(netloc, 0)
        delay = random.uniform(self.min_delay, self.max_delay)
        to_sleep = max(0.0, (last + delay) - now)
        if to_sleep > 0:
            time.sleep(to_sleep)
        self.last_hit[netloc] = time.time()


def fetch_url(session: requests.Session, limiter: RateLimiter, url: str) -> Tuple[Optional[str], Optional[str], Optional[int], Optional[int]]:
    try:
        limiter.wait(url)
        resp = session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        text = resp.text if resp.content else ""
        size_bytes = len(resp.content or b"")
        return text, resp.url, resp.status_code, size_bytes
    except requests.RequestException:
        return None, None, None, None


def build_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    })
    return s


def clean_visible_text(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    for tag in [
        "script",
        "style",
        "noscript",
        "svg",
        "canvas",
        "form",
        "input",
        "button",
        "header",
        "footer",
        "nav",
        "aside",
    ]:
        for el in soup.find_all(tag):
            el.decompose()
    for el in soup(text=lambda t: isinstance(t, Comment)):
        el.extract()
    for el in soup.find_all(attrs={"aria-hidden": True}):
        el.decompose()
    for el in soup.find_all(style=True):
        style = el.get("style", "").lower()
        if "display:none" in style or "visibility:hidden" in style:
            el.decompose()
    text = soup.get_text(separator=" ", strip=True)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def discover_about_links(base_html: str, base_url: str, resolved_base: str) -> List[str]:
    from urllib.parse import urljoin, urlparse

    soup = BeautifulSoup(base_html, "html.parser")
    candidates = []
    about_keywords = [
        "about",
        "info",
        "bio",
        "statement",
        "profile",
        "cv",
    ]
    base_netloc = urlparse(resolved_base or base_url).netloc
    seen = set()
    for a in soup.find_all("a", href=True):
        href = a.get("href")
        if not href:
            continue
        label = (a.get_text(" ", strip=True) or "").lower()
        full = urljoin(resolved_base or base_url, href)
        if full in seen:
            continue
        seen.add(full)
        p = urlparse(full)
        if p.netloc and p.netloc != base_netloc:
            continue
        path_lower = (p.path or "").lower()
        match = any(k in path_lower for k in about_keywords) or any(
            k in label for k in about_keywords
        )
        if match:
            candidates.append(full)
    # de-duplicate preserving order
    out = []
    s = set()
    for u in candidates:
        if u not in s:
            s.add(u)
            out.append(u)
    return out[:MAX_ABOUT_PAGES]


def extract_education_excerpt(text: str) -> str:
    if not text:
        return ""
    lower = text.lower()
    if not any(term in lower for term in EDU_TERMS):
        return ""
    parts = re.split(r"(?<=[\.!?])\s+", text)
    hits = [p for p in parts if any(re.search(rf"\b{term}\b", p, re.I) for term in EDU_TERMS)]
    snippet = " ".join(hits)
    return snippet[:500]


def compute_booleans(text: str) -> Tuple[bool, bool]:
    if not text:
        return False, False
    t = text.lower()
    # mentions_design
    design_terms = [
        r"\bdesign\b",
        r"\bdesigner\b",
        r"graphic design",
        r"communication design",
        r"visual design",
        r"design studio",
    ]
    mentions_design = any(re.search(pat, t, re.I) for pat in design_terms)
    # mentions_art
    art_terms = [
        r"\bart\b",
        r"\bartist\b",
        r"\bartistic\b",
        r"contemporary art",
        r"visual art",
        r"media art",
    ]
    mentions_art = any(re.search(pat, t, re.I) for pat in art_terms)
    return mentions_design, mentions_art


def tokenize(text: str) -> List[str]:
    tokens = re.findall(r"[a-zA-Z][a-zA-Z\-]{2,}", text.lower())
    cleaned = []
    for tok in tokens:
        tok = tok.strip("-_")
        if len(tok) < 3:
            continue
        if tok in STOPWORDS:
            continue
        cleaned.append(tok)
    return cleaned


def extract_keywords(text: str, top_n: int = 10) -> List[str]:
    if not text:
        return []
    toks = tokenize(text)
    if not toks:
        return []
    wl = set(WHITELIST_KEYWORDS)
    counts = Counter(toks)
    whitelisted_present = [w for w in WHITELIST_KEYWORDS if w in counts]
    whitelisted_present.sort(key=lambda w: (-counts[w], w))
    keywords = []
    for w in whitelisted_present:
        if w not in keywords:
            keywords.append(w)
        if len(keywords) >= top_n:
            return keywords[:top_n]
    # fill with most frequent remaining tokens
    for tok, _ in counts.most_common(100):
        if tok in keywords:
            continue
        if tok in wl:
            continue
        keywords.append(tok)
        if len(keywords) >= top_n:
            break
    return keywords[:top_n]


def classify(mentions_design: bool, mentions_art: bool, keywords: List[str], text_length: int, pages_count: int) -> Tuple[str, str]:
    if text_length < 150 or pages_count == 0:
        return "Unclear_missing", "Insufficient text or no about page"
    kw_set = set(keywords)
    if mentions_art and (kw_set & INTERDISCIPLINARY_HINTS):
        return "Interdisciplinary", "mentions_art and keywords include interdisciplinary terms"
    if mentions_design and not mentions_art:
        gcount = len(kw_set & GRAPHIC_KEYWORDS)
        if gcount >= max(3, len(keywords) // 2):
            return "Graphic-only", "mentions_design, mostly graphic keywords"
    return "Unclear_missing", "Ambiguous signals"


def confidence_score(text_length: int, pages_count: int, has_education: bool) -> int:
    score = 0
    score += min(text_length, 1500) / 1500 * 60
    score += min(pages_count, 4) / 4 * 25
    if has_education:
        score += 15
    return int(max(0, min(100, round(score))))


def write_excel(df: pd.DataFrame, path: str):
    sheets = ["Interdisciplinary", "Graphic-only", "Unclear_missing"]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for s in sheets:
            sub = df[df["category"] == s]
            sub.to_excel(writer, sheet_name=s, index=False)
    wb = load_workbook(path)
    for s in sheets:
        ws = wb[s]
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
        # Optional: modest column width pass
        for c in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(c)].width = 18
    wb.save(path)


def process_row(
    session: requests.Session,
    limiter: RateLimiter,
    logger: RunLogger,
    name: str,
    url: Optional[str],
    notes: Optional[str],
) -> Dict[str, object]:
    base = {
        "name": name,
        "url": url or "",
        "resolved_url": "",
        "source_pages": "",
        "status": "",
        "about_text_excerpt": "",
        "text_length": 0,
        "education_excerpt": "",
        "mentions_design": False,
        "mentions_art": False,
        "self_keywords": [],
        "category": "Unclear_missing",
        "category_reason": "",
        "confidence_score": 0,
    }

    if not url or not url.strip():
        base["status"] = "missing_url"
        logger.log(name, url, "row", "missing_url")
        return base

    input_url = normalize_url(url)
    logger.log(name, url, "fetch_homepage", "start", page_url=input_url)
    html, resolved, http_status, size_bytes = fetch_url(session, limiter, input_url)
    if resolved:
        base["resolved_url"] = resolved
    if html is None or not http_status or http_status >= 400:
        base["status"] = f"fetch_error"
        logger.log(
            name,
            url,
            "fetch_homepage",
            "error",
            page_url=input_url,
            http_status=http_status or 0,
            size_bytes=size_bytes or 0,
        )
        return base
    logger.log(
        name,
        url,
        "fetch_homepage",
        "ok",
        page_url=resolved,
        http_status=http_status,
        size_bytes=size_bytes,
    )

    home_text = clean_visible_text(html) if html else ""
    about_links = discover_about_links(html or "", input_url, resolved or input_url)

    pages_fetched: List[str] = []
    combined_text_parts: List[str] = []

    if home_text:
        combined_text_parts.append(home_text)
        pages_fetched.append(resolved or input_url)

    for link in about_links[:MAX_ABOUT_PAGES]:
        logger.log(name, url, "fetch_about", "start", page_url=link)
        ahtml, ares, astatus, asize = fetch_url(session, limiter, link)
        if ahtml is None or not astatus or astatus >= 400:
            logger.log(
                name,
                url,
                "fetch_about",
                "error",
                page_url=link,
                http_status=astatus or 0,
                size_bytes=asize or 0,
            )
            continue
        logger.log(
            name,
            url,
            "fetch_about",
            "ok",
            page_url=ares or link,
            http_status=astatus,
            size_bytes=asize,
        )
        text = clean_visible_text(ahtml)
        if text:
            combined_text_parts.append(text)
            pages_fetched.append(ares or link)

    combined_text = "\n\n".join([p for p in combined_text_parts if p])
    text_length = len(combined_text)
    excerpt = combined_text[:MAX_EXCERPT_CHARS]

    education = extract_education_excerpt(combined_text)
    mentions_design, mentions_art = compute_booleans(combined_text)
    keywords = extract_keywords(combined_text)
    category, reason = classify(
        mentions_design, mentions_art, keywords, text_length, len(pages_fetched)
    )
    conf = confidence_score(text_length, len(pages_fetched), bool(education))

    base.update(
        {
            "source_pages": "; ".join(pages_fetched[: 1 + MAX_ABOUT_PAGES]),
            "status": "ok" if text_length > 0 else "no_text",
            "about_text_excerpt": excerpt,
            "text_length": text_length,
            "education_excerpt": education,
            "mentions_design": bool(mentions_design),
            "mentions_art": bool(mentions_art),
            "self_keywords": ", ".join(keywords),
            "category": category,
            "category_reason": reason,
            "confidence_score": conf,
        }
    )
    logger.log(name, url, "classify", "ok", detail=f"{category}: {reason}")
    return base


def run(input_csv: str, output_xlsx: str, log_csv: str):
    try:
        df = pd.read_csv(input_csv)
    except Exception as e:
        print(f"Failed to read input CSV: {e}", file=sys.stderr)
        sys.exit(1)

    req_cols = {"name", "url"}
    missing = req_cols - set(df.columns.str.lower())
    # Try case-insensitive access
    columns_lower = {c.lower(): c for c in df.columns}
    if missing:
        # Verify again with direct check
        if not ("name" in columns_lower and "url" in columns_lower):
            print("Input CSV must contain columns: name, url", file=sys.stderr)
            sys.exit(1)

    name_col = columns_lower.get("name", "name")
    url_col = columns_lower.get("url", "url")
    notes_col = columns_lower.get("notes")

    session = build_session()
    limiter = RateLimiter(1.0, 2.0)
    logger = RunLogger(log_csv)

    out_rows: List[Dict[str, object]] = []
    for idx, row in df.iterrows():
        try:
            name = str(row.get(name_col, "") or "").strip()
            url = row.get(url_col)
            notes = str(row.get(notes_col, "") or "").strip() if notes_col else ""
            res = process_row(session, limiter, logger, name, url, notes)
            out_rows.append(res)
        except Exception as e:
            logger.log(str(row.get(name_col, "")), str(row.get(url_col, "")), "row", "exception", detail=str(e))
            out_rows.append(
                {
                    "name": str(row.get(name_col, "") or "").strip(),
                    "url": str(row.get(url_col, "") or "").strip(),
                    "resolved_url": "",
                    "source_pages": "",
                    "status": "exception",
                    "about_text_excerpt": "",
                    "text_length": 0,
                    "education_excerpt": "",
                    "mentions_design": False,
                    "mentions_art": False,
                    "self_keywords": "",
                    "category": "Unclear_missing",
                    "category_reason": "Processing error",
                    "confidence_score": 0,
                }
            )

    out_df = pd.DataFrame(out_rows)
    write_excel(out_df, output_xlsx)
    print(f"Wrote {output_xlsx} with {len(out_df)} rows across 3 sheets.")
    print(f"Run log available at {log_csv}.")


def parse_args(argv: Optional[List[str]] = None):
    p = argparse.ArgumentParser(description="Enrich artist/studio websites via polite crawling.")
    p.add_argument("--input", default="input.csv", help="Input CSV path (columns: name,url,notes)")
    p.add_argument("--output", default="artists_studios_enrichment.xlsx", help="Output Excel path")
    p.add_argument("--log", default="run_log.csv", help="Run log CSV path")
    return p.parse_args(argv)


if __name__ == "__main__":
    args = parse_args()
    run(args.input, args.output, args.log)
